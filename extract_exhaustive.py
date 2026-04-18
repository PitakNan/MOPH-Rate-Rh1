import openpyxl
import json
import os

def extract_units_exhaustive():
    file_path = '0 ผลการตรวจสอบรายการอัตราค่าบริการสาธารณสุขฯ.xlsx'
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        unique_units = {} # Keyed by normalized name
        
        # Priority mapping from sheet 2 (ࢵ 1) which has more details
        s2_name = wb.sheetnames[1] if len(wb.sheetnames) > 1 else None
        if s2_name:
            print(f"Processing detail sheet: {s2_name}")
            sheet = wb[s2_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 7: continue
                name = str(row[3]).strip() if row[3] else ""
                if not name or name == 'None' or 'หน่วยงาน' in name: continue
                
                norm_name = name.replace(" ", "").lower()
                
                # Check for rate details (Cols 11-15)
                try:
                    r_total = int(row[11]) if row[11] is not None and str(row[11]).isdigit() else 0
                    r_thai_equal = int(row[12]) if row[12] is not None and str(row[12]).isdigit() else 0
                    r_thai_lower = int(row[13]) if row[13] is not None and str(row[13]).isdigit() else 0
                    r_thai_higher = int(row[14]) if row[14] is not None and str(row[14]).isdigit() else 0
                    r_inter_higher = int(row[15]) if row[15] is not None and str(row[15]).isdigit() else 0
                except:
                    r_total = r_thai_equal = r_thai_lower = r_thai_higher = r_inter_higher = 0

                unique_units[norm_name] = {
                    "code": str(row[6]).strip() if row[6] else "",
                    "province": str(row[4]).strip() if row[4] else "",
                    "name": name,
                    "type": str(row[5]).strip() if row[5] else "",
                    "status": 0,
                    "startDate": str(row[7]).split(' ')[0] if row[7] and ' ' in str(row[7]) else str(row[7]),
                    "outcome": "",
                    "updateDate": "",
                    "fileLink": "",
                    "rate_total": r_total,
                    "rate_thai_lower": r_thai_lower,
                    "rate_inter_lower": 0,
                    "rate_thai_equal": r_thai_equal,
                    "rate_inter_equal": 0,
                    "rate_thai_higher": r_thai_higher,
                    "rate_inter_higher": r_inter_higher
                }

        # Base mapping from sheet 1 (หน่วยบริการ - index 0)
        s1_name = wb.sheetnames[0]
        print(f"Processing base sheet: {s1_name}")
        sheet = wb[s1_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3: continue
            name = str(row[2]).strip() if row[2] else ""
            if not name or name == 'None' or 'หน่วยงาน' in name: continue
            
            norm_name = name.replace(" ", "").lower()
            
            # Unit Code is Column B (index 1) in S1
            code_val = str(row[1]).strip() if row[1] else ""
            
            if norm_name not in unique_units:
                unique_units[norm_name] = {
                    "code": code_val,
                    "province": str(row[0]).strip() if row[0] else "",
                    "name": name,
                    "type": str(row[3]).strip() if row[3] else "",
                    "status": 0,
                    "startDate": "",
                    "outcome": "",
                    "updateDate": "",
                    "fileLink": "",
                    "rate_total": 0,
                    "rate_thai_lower": 0,
                    "rate_inter_lower": 0,
                    "rate_thai_equal": 0,
                    "rate_inter_equal": 0,
                    "rate_thai_higher": 0,
                    "rate_inter_higher": 0
                }
            else:
                # Always trust Sheet 1 for code if it matches name
                if code_val:
                    unique_units[norm_name]["code"] = code_val
            
        final_list = list(unique_units.values())
        
        with open('dashboard_moph/data.js', 'w', encoding='utf-8') as f:
            f.write('const initialData = ' + json.dumps(final_list, ensure_ascii=False, indent=2) + ';')
        
        with open('dashboard_moph/data.json', 'w', encoding='utf-8') as f:
            json.dump(final_list, f, ensure_ascii=False, indent=2)
        
        print(f"Successfully extracted {len(final_list)} unique units.")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    extract_units_exhaustive()
