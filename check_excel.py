import openpyxl
try:
    wb = openpyxl.load_workbook('0 ผลการตรวจสอบรายการอัตราค่าบริการสาธารณสุขฯ.xlsx', read_only=True)
    print("Sheets found:", wb.sheetnames)
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"Sheet '{name}' has {sheet.max_row} rows.")
except Exception as e:
    print("Error:", e)
